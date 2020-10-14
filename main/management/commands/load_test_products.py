from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from main.models import Product, Operation


def action(count=None, file_path=None):
    if not file_path:
        file_path = str(settings.BASE_DIR) + '/test_data/products.xlsx'
    work_book = load_workbook(file_path)
    sheet = work_book.get_sheet_by_name('data')

    data = []
    row = 2
    while True:
        if count is not None and row > count:
            break
        title = sheet.cell(row=row, column=1).value
        price = sheet.cell(row=row, column=2).value
        if not title:
            break
        data.append(Product(title=title, price=price))
        row += 1

    Product.objects.bulk_create(data)

    return row - 2


class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument(
            '--count',
            type=int,
            help='Количество создаваемых в базе записей'
        )

    def handle(self, *args, **options):
        Product.objects.all().delete()
        count = options['count']
        created_records = action(count)
        Operation.objects.create(
            username='- Администратор системы -',
            operation=f'Командой load_test_products создано {created_records} записей о товарах'
        )
        print(f'Создано записей: {created_records}')
