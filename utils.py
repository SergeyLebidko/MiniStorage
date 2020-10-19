import os
from io import BytesIO
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from main.models import DocumentItem, StorageItem


def get_tmp_folder_path():
    return str(settings.BASE_DIR) + settings.TMP_FOLDER


def get_tmp_file_path(filename):
    return get_tmp_folder_path() + filename


def check_tmp_folder():
    tmp_folder = get_tmp_folder_path()
    try:
        os.mkdir(tmp_folder)
    except OSError:
        pass


def model_to_xls(model, column_descriptions):
    data = model.objects.all()
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = 'data'

    # Создаем шапку таблицы
    for column, description in enumerate(column_descriptions, 1):
        cell = work_sheet.cell(row=1, column=column, value=description['display_name'])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        width = description.get('width')
        if width:
            work_sheet.column_dimensions[get_column_letter(column)].width = width

    # Копируем данные из БД
    row = 2
    for data_element in data:
        for column, description in enumerate(column_descriptions, 1):
            value = getattr(data_element, description['machine_name'])
            subs = description.get('subs')
            if subs:
                value = subs[value]
            work_sheet.cell(row=row, column=column, value=value)
        row += 1

    # Сохраняем данные во временный файл, затем отдаем содержимое файла клиенту. Сам файл удаляем
    check_tmp_folder()
    tmp_file_path = get_tmp_file_path('products.xlsx')
    work_book.save(tmp_file_path)

    with open(tmp_file_path, 'rb') as file:
        bio = BytesIO(file.read())
    os.remove(tmp_file_path)

    return bio


def get_username_for_operation(user):
    return user.get_full_name() or user.get_username()


def apply_receipt_document(document):
    document_items = DocumentItem.objects.filter(document=document)
    for document_item in document_items:
        storage_item = StorageItem.objects.filter(product_id=document_item.product_id).first()
        if storage_item:
            storage_item.count += document_item.count
            storage_item.save()
        else:
            StorageItem.objects.create(product=document_item.product, count=document_item.count)


def apply_expense_document(document):
    document_items = DocumentItem.objects.filter(document=document)
    storage_items = []
    for document_item in document_items:
        storage_item = StorageItem.objects.filter(product_id=document_item.product_id).first()
        if not storage_item:
            raise Exception(document_item.product.title)
        storage_item.count -= document_item.count
        if storage_item.count < 0:
            raise Exception(document_item.product.title)
        storage_items.append(storage_item)

    for storage_item in storage_items:
        if storage_item.count == 0:
            storage_item.delete()
        else:
            storage_item.save()


def unapply_receipt_document(document):
    apply_expense_document(document)


def unapply_expense_document(document):
    apply_receipt_document(document)
